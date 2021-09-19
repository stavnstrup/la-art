from collections import OrderedDict
from itertools import islice
from os import mkdir
from openpyxl import load_workbook
import os

# load the workbook
wb = load_workbook('LA-ART-data.xlsx')
sheet = wb['Ark1']

# list to hold art
art_list = []

# Iterate through each row in worksheet and fetch values into dict

for row in islice(sheet.values, 1, sheet.max_row):
    art = OrderedDict()
    art['udstilles'] = row[0]
    art['title'] = row[1]
    art['height'] = row[2]
    art['width'] = row[3]
    art['method'] = row[4]
    art['year'] = row[5]
    art['price'] = row[6]
    art['rasterWidth'] = int(row[7])
    art['rasterHeight'] = int(row[8])
    art['fileName'] = row[9]
    art['medie'] = row[10]
    art['show'] = row[11]
    art['slug'] = row[12]
    art['weight'] = int(row[13])
    art_list.append(art)

os.mkdir('maleri')
os.mkdir('tegning')

for art in art_list:
    os.mkdir(art['medie'] + '/' + art['slug'])
    f = open('{0}/{1}/index.md'.format(art['medie'], art['slug']), 'w')
    f.write("---\n")
    for (key, value) in art.items():
        if key in ['year', 'rasterWidth', 'rasterHeight']:
            f.write("{0}: {1}\n".format(key, value))
        else:
            f.write("{0}: \"{1}\"\n".format(key, value))
    f.write("---\n")
    f.close()

f = open('fiximages.sh', 'w')
f.write('#!/bin/sh' + "\n")
for art in art_list:
    f.write(('mv {0}/{2} {0}/{1}/' +
            "\n").format(art['medie'], art['slug'], art['fileName']))
f.close()
